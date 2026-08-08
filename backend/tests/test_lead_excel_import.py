from io import BytesIO

from openpyxl import Workbook, load_workbook
import xlwt

from app.modules.leads import excel_import
from tests.conftest import auth_header


def _row(
    code: str,
    name: str,
    phone: str,
    *,
    school: str = "实验学校",
    grade: str = "三年级",
    age: object = 9,
    source: str = "微信",
    campus: str = "中心校区",
    owner: str = "运营",
    creator: str = "历史创建人",
    notes: str = "待联系",
) -> list[object]:
    return [code, name, phone, school, grade, age, source, campus, owner, creator, notes]


def _xlsx(rows: list[list[object]], headers: list[str] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers or excel_import.HEADERS)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xls(rows: list[list[object]]) -> bytes:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("线索导入")
    for column, header in enumerate(excel_import.HEADERS):
        sheet.write(0, column, header)
    for row_index, row in enumerate(rows, start=1):
        for column, value in enumerate(row):
            sheet.write(row_index, column, value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _upload(client, headers: dict, content: bytes, filename: str):
    return client.post(
        "/api/v1/leads/import",
        headers=headers,
        files={"file": (filename, content, "application/octet-stream")},
    )


def test_download_import_template(client):
    headers = auth_header(client, "ops", "ops123")
    response = client.get("/api/v1/leads/import-template", headers=headers)
    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert [cell.value for cell in workbook.active[1]] == excel_import.HEADERS


def test_import_xlsx_maps_fields_source_owner_and_activity(client):
    headers = auth_header(client, "ops", "ops123")
    content = _xlsx(
        [
            _row("L-001", "张同学", "13800002001", source="转介绍"),
            _row("L-002", "李同学", "13800002002", source="小红书", owner=""),
        ]
    )
    response = _upload(client, headers, content, "leads.xlsx")
    assert response.status_code == 200, response.text
    result = response.json()["data"]
    assert result == {
        "imported_count": 2,
        "duplicate_count": 0,
        "failed_count": 0,
        "warning_count": 0,
        "details": [
            {"row": 2, "status": "imported", "message": "导入成功"},
            {"row": 3, "status": "imported", "message": "导入成功"},
        ],
    }

    listed = client.get("/api/v1/leads", headers=headers, params={"page_size": 100})
    items = {item["external_code"]: item for item in listed.json()["data"]["items"]}
    first = items["L-001"]
    assert first["school"] == "实验学校"
    assert first["grade"] == "三年级"
    assert first["age"] == 9
    assert first["campus"] == "中心校区"
    assert first["imported_creator_name"] == "历史创建人"
    assert first["source"] == "referral"
    assert first["owner_name"] == "运营"
    second = items["L-002"]
    assert second["source"] == "other"
    assert second["channel_note"] == "原来源渠道：小红书"

    activities = client.get(f"/api/v1/leads/{first['id']}/activities", headers=headers)
    create_activity = activities.json()["data"]["items"][0]
    assert create_activity["kind"] == "create"
    assert create_activity["actor_name"] == "运营"


def test_import_legacy_xls(client):
    headers = auth_header(client, "ops", "ops123")
    response = _upload(
        client,
        headers,
        _xls([_row("XLS-001", "旧表线索", "13800002003", source="大众点评")]),
        "legacy.xls",
    )
    assert response.status_code == 200, response.text
    assert response.json()["data"]["imported_count"] == 1
    listed = client.get("/api/v1/leads", headers=headers, params={"name": "旧表线索"})
    item = listed.json()["data"]["items"][0]
    assert item["external_code"] == "XLS-001"
    assert item["source"] == "dianping"


def test_import_partial_failures_duplicates_and_owner_warning(client):
    headers = auth_header(client, "ops", "ops123")
    first = _upload(
        client,
        headers,
        _xlsx([_row("DB-001", "已有线索", "13800002004")]),
        "first.xlsx",
    )
    assert first.status_code == 200

    rows = [
        _row("DB-002", "手机号重复", "13800002004"),
        _row("DB-001", "编号重复", "13800002005"),
        _row("NEW-001", "文件首条", "13800002006"),
        _row("NEW-002", "文件手机号重复", "13800002006"),
        _row("NEW-001", "文件编号重复", "13800002007"),
        _row("BAD-001", "", "13800002008"),
        _row("BAD-002", "坏手机", "123"),
        _row("BAD-003", "坏年龄", "13800002009", age=100),
        _row("WARN-001", "未分配线索", "13800002010", owner="不存在的人"),
    ]
    response = _upload(client, headers, _xlsx(rows), "mixed.xlsx")
    assert response.status_code == 200, response.text
    result = response.json()["data"]
    assert result["imported_count"] == 2
    assert result["duplicate_count"] == 4
    assert result["failed_count"] == 3
    assert result["warning_count"] == 1
    assert any(item["status"] == "warning" and item["row"] == 10 for item in result["details"])

    warning_lead = client.get(
        "/api/v1/leads", headers=headers, params={"name": "未分配线索"}
    ).json()["data"]["items"][0]
    assert warning_lead["owner_id"] is None


def test_invalid_row_does_not_reserve_duplicate_values(client):
    headers = auth_header(client, "ops", "ops123")
    rows = [
        _row("SAME", "", "13800002011"),
        _row("SAME", "有效线索", "13800002011"),
    ]
    result = _upload(client, headers, _xlsx(rows), "priority.xlsx").json()["data"]
    assert result["failed_count"] == 1
    assert result["imported_count"] == 1
    assert result["duplicate_count"] == 0


def test_import_rejects_invalid_files_and_limits(client):
    headers = auth_header(client, "ops", "ops123")
    unsupported = _upload(client, headers, b"not excel", "leads.csv")
    assert unsupported.status_code == 400

    too_large = _upload(
        client,
        headers,
        b"x" * (excel_import.MAX_FILE_SIZE + 1),
        "large.xlsx",
    )
    assert too_large.status_code == 400
    assert "5 MB" in too_large.json()["error"]["message"]

    missing_headers = excel_import.HEADERS.copy()
    missing_headers.remove("手机号")
    missing = _upload(client, headers, _xlsx([], missing_headers), "missing.xlsx")
    assert missing.status_code == 400
    assert "手机号" in missing.json()["error"]["message"]

    too_many_rows = [_row(f"ROW-{index}", f"线索{index}", f"138{index:08d}") for index in range(1001)]
    too_many = _upload(client, headers, _xlsx(too_many_rows), "too-many.xlsx")
    assert too_many.status_code == 400
    assert "1000" in too_many.json()["error"]["message"]


def test_import_requires_leads_write(client):
    headers = auth_header(client, "teacher1", "t123")
    response = _upload(
        client,
        headers,
        _xlsx([_row("NOAUTH", "无权限", "13800002012")]),
        "forbidden.xlsx",
    )
    assert response.status_code == 403


def test_database_error_rolls_back_entire_import(client, monkeypatch):
    headers = auth_header(client, "ops", "ops123")
    original_add_activity = excel_import.add_activity
    calls = 0

    def fail_on_second(*args, **kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("simulated database failure")
        return original_add_activity(*args, **kwargs)

    monkeypatch.setattr(excel_import, "add_activity", fail_on_second)
    response = _upload(
        client,
        headers,
        _xlsx(
            [
                _row("ROLL-001", "回滚一", "13800002013"),
                _row("ROLL-002", "回滚二", "13800002014"),
            ]
        ),
        "rollback.xlsx",
    )
    assert response.status_code == 500
    listed = client.get("/api/v1/leads", headers=headers, params={"name": "回滚"})
    assert listed.json()["data"]["total"] == 0
