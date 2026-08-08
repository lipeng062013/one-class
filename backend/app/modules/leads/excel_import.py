from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from sqlalchemy.orm import Session

from app.models.lead import Lead
from app.models.user import User
from app.modules.leads.service import add_activity

MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_DATA_ROWS = 1000
HEADERS = [
    "编号",
    "姓名",
    "手机号",
    "学校",
    "年级",
    "年龄",
    "来源渠道",
    "对应校区",
    "归属人",
    "创建人",
    "备注",
]

SOURCE_MAP = {
    "老带新": "referral",
    "转介绍": "referral",
    "大众点评": "dianping",
    "微信": "wechat",
    "到店": "walkin",
    "其他": "other",
    "": "other",
}
PHONE_PATTERN = re.compile(r"^1\d{10}$")


class ImportFileError(ValueError):
    pass


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_phone(value: Any) -> str:
    return re.sub(r"\s+", "", _cell_text(value))


def _parse_age(value: Any) -> tuple[int | None, str | None]:
    text = _cell_text(value)
    if not text:
        return None, None
    try:
        number = float(text)
    except ValueError:
        return None, "年龄必须是 1 至 99 的整数"
    if not number.is_integer() or not 1 <= number <= 99:
        return None, "年龄必须是 1 至 99 的整数"
    return int(number), None


def _validate_headers(values: list[Any]) -> None:
    actual = [_cell_text(value) for value in values[: len(HEADERS)]]
    if actual != HEADERS:
        missing = [header for header in HEADERS if header not in actual]
        if missing:
            raise ImportFileError(f"缺少必需表头：{'、'.join(missing)}")
        raise ImportFileError("表头顺序不正确，请下载并使用系统导入模板")


def _row_payload(values: list[Any]) -> dict[str, Any]:
    padded = list(values[: len(HEADERS)]) + [None] * max(0, len(HEADERS) - len(values))
    return dict(zip(HEADERS, padded, strict=True))


def _parse_xlsx(content: bytes) -> list[tuple[int, dict[str, Any]]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ImportFileError("工作簿为空")
        _validate_headers(list(header))
        parsed = [
            (row_number, _row_payload(list(values)))
            for row_number, values in enumerate(rows, start=2)
            if any(_cell_text(value) for value in values)
        ]
        workbook.close()
        return parsed
    except ImportFileError:
        raise
    except Exception as exc:
        raise ImportFileError("无法解析 .xlsx 文件，请确认文件未损坏") from exc


def _parse_xls(content: bytes) -> list[tuple[int, dict[str, Any]]]:
    import xlrd

    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        if workbook.nsheets < 1:
            raise ImportFileError("工作簿为空")
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows < 1:
            raise ImportFileError("工作簿为空")
        _validate_headers(sheet.row_values(0))
        parsed = [
            (index + 1, _row_payload(sheet.row_values(index)))
            for index in range(1, sheet.nrows)
            if any(_cell_text(value) for value in sheet.row_values(index))
        ]
        workbook.release_resources()
        return parsed
    except ImportFileError:
        raise
    except Exception as exc:
        raise ImportFileError("无法解析 .xls 文件，请确认文件未损坏") from exc


def parse_workbook(content: bytes, extension: str) -> list[tuple[int, dict[str, Any]]]:
    if extension == ".xlsx":
        rows = _parse_xlsx(content)
    elif extension == ".xls":
        rows = _parse_xls(content)
    else:
        raise ImportFileError("仅支持 .xls 或 .xlsx 文件")
    if len(rows) > MAX_DATA_ROWS:
        raise ImportFileError(f"数据行数不能超过 {MAX_DATA_ROWS} 条")
    return rows


def build_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "线索导入"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="A16207")
        cell.alignment = Alignment(horizontal="center")
    widths = [16, 14, 16, 22, 12, 10, 16, 18, 14, 14, 30]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _owner_index(db: Session) -> dict[str, set[int]]:
    users = (
        db.query(User)
        .filter(User.is_active.is_(True), User.deleted_at.is_(None))
        .all()
    )
    index: dict[str, set[int]] = {}
    for user in users:
        for value in {user.display_name.strip(), user.username.strip()}:
            if value:
                index.setdefault(value, set()).add(user.id)
    return index


def _existing_values(db: Session) -> tuple[set[str], set[str]]:
    phones = {
        value.strip()
        for (value,) in db.query(Lead.phone).filter(Lead.phone.is_not(None)).all()
        if value and value.strip()
    }
    codes = {
        value.strip()
        for (value,) in db.query(Lead.external_code)
        .filter(Lead.external_code.is_not(None))
        .all()
        if value and value.strip()
    }
    return phones, codes


def import_rows(
    db: Session,
    rows: list[tuple[int, dict[str, Any]]],
    actor: User,
) -> dict[str, Any]:
    owner_index = _owner_index(db)
    existing_phones, existing_codes = _existing_values(db)
    seen_phones: set[str] = set()
    seen_codes: set[str] = set()
    imported_count = duplicate_count = failed_count = warning_count = 0
    details: list[dict[str, Any]] = []

    try:
        for row_number, row in rows:
            name = _cell_text(row["姓名"])
            phone = _normalize_phone(row["手机号"])
            external_code = _cell_text(row["编号"])
            age, age_error = _parse_age(row["年龄"])

            errors: list[str] = []
            if not name:
                errors.append("姓名不能为空")
            if not phone:
                errors.append("手机号不能为空")
            elif PHONE_PATTERN.fullmatch(phone) is None:
                errors.append("手机号必须为 11 位中国大陆手机号")
            if age_error:
                errors.append(age_error)
            if errors:
                failed_count += 1
                details.append({"row": row_number, "status": "failed", "message": "；".join(errors)})
                continue

            duplicate_reasons: list[str] = []
            if phone in seen_phones:
                duplicate_reasons.append("手机号在当前文件中重复")
            elif phone in existing_phones:
                duplicate_reasons.append("手机号已存在")
            if external_code:
                if external_code in seen_codes:
                    duplicate_reasons.append("编号在当前文件中重复")
                elif external_code in existing_codes:
                    duplicate_reasons.append("编号已存在")
            seen_phones.add(phone)
            if external_code:
                seen_codes.add(external_code)
            if duplicate_reasons:
                duplicate_count += 1
                details.append(
                    {"row": row_number, "status": "duplicate", "message": "；".join(duplicate_reasons)}
                )
                continue

            source_text = _cell_text(row["来源渠道"])
            source = SOURCE_MAP.get(source_text, "other")
            channel_note = "" if source_text in SOURCE_MAP else f"原来源渠道：{source_text}"
            owner_text = _cell_text(row["归属人"])
            owner_ids = owner_index.get(owner_text, set()) if owner_text else set()
            owner_id = next(iter(owner_ids)) if len(owner_ids) == 1 else None
            warnings: list[str] = []
            if owner_text and not owner_ids:
                warnings.append(f"归属人“{owner_text}”未匹配，已留空")
            elif len(owner_ids) > 1:
                warnings.append(f"归属人“{owner_text}”匹配到多个用户，已留空")

            lead = Lead(
                external_code=external_code or None,
                student_or_parent_name=name,
                phone=phone,
                school=_cell_text(row["学校"]),
                grade=_cell_text(row["年级"]),
                age=age,
                source=source,
                channel_note=channel_note,
                campus=_cell_text(row["对应校区"]),
                owner_id=owner_id,
                imported_creator_name=_cell_text(row["创建人"]),
                notes=_cell_text(row["备注"]),
                status="new",
            )
            db.add(lead)
            db.flush()
            add_activity(
                db,
                lead_id=lead.id,
                actor=actor,
                kind="create",
                title="导入线索",
                content=f"通过 Excel 导入线索「{lead.student_or_parent_name}」",
                meta={"source": lead.source, "external_code": lead.external_code},
            )
            imported_count += 1
            existing_phones.add(phone)
            if external_code:
                existing_codes.add(external_code)
            if warnings:
                warning_count += 1
                details.append(
                    {"row": row_number, "status": "warning", "message": "；".join(warnings)}
                )
            else:
                details.append({"row": row_number, "status": "imported", "message": "导入成功"})

        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "imported_count": imported_count,
        "duplicate_count": duplicate_count,
        "failed_count": failed_count,
        "warning_count": warning_count,
        "details": details,
    }
